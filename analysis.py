import streamlit as st
import statistics
import base64
import io
import os
import matplotlib.pyplot as plt
import numpy as np
from tabulate import tabulate
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenPyXLImage
from PIL import Image as PILImage
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenPyXLImage
from PIL import Image as PILImage
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment
from import_export import generate_settings_xlsx
from scipy.stats import t


def generate_analysis(chat_results, analyze_rating=True, analyze_length=True):
    total_cost = sum(chat_res["total_cost"] for chat_res in chat_results.values())

    # Prepare the analysis data table
    headers = ["Metric"] + [f"Chat {chat_index}" for chat_index in sorted(chat_results.keys())]
    analysis_data = [headers]

    if analyze_length:
        lengths_data = []
        for chat_index in sorted(chat_results.keys()):
            # Filter out None values from lengths
            lengths = [l for l in chat_results[chat_index]["lengths"] if l is not None]
            lengths_data.append(lengths)

        metrics = ["Average Length", "Median Length", "Std Dev Length", "Min Length", "Max Length"]
        stats_funcs = [statistics.mean, statistics.median, statistics.stdev, min, max]
        for metric, func in zip(metrics, stats_funcs):
            row = [metric]
            for lengths in lengths_data:
                if lengths:
                    if func == statistics.stdev and len(lengths) < 2:
                        value = "N/A"
                    else:
                        value = f"{func(lengths):.1f}"
                else:
                    value = "N/A"
                row.append(value)
            analysis_data.append(row)

    row = ["Iterations per chat"]
    for chat_index in sorted(chat_results.keys()):
        iterations = len(chat_results[chat_index]["responses"])
        row.append(str(iterations))
    analysis_data.append(row)

    if analyze_rating:
        ratings_data = []
        empty_rating_chats = []  # To track chats with no valid ratings
        for chat_index in sorted(chat_results.keys()):
            # Filter out None values from ratings
            ratings = [r for r in chat_results[chat_index]["ratings"] if r is not None]
            if not ratings:
                empty_rating_chats.append(f"Chat {chat_index}")
            ratings_data.append(ratings)

        # Print error if any chat has no valid ratings
        if empty_rating_chats:
            error_message = f"Error: No valid ratings found for {', '.join(empty_rating_chats)}."
            print(error_message)  # Prints to console
            st.error(error_message)  # Displays error in Streamlit app

        metrics = ["Average Rating", "95% CI of Rating", "SEM of Rating", "Std Dev of Rating", "Minimum Rating", "Maximum Rating"]
        stats_funcs = [
            statistics.mean,
            lambda x: f"±{1.96 * statistics.stdev(x) / np.sqrt(len(x)):.2f}",  # 95% CI normal-based
            lambda x: statistics.stdev(x) / np.sqrt(len(x)),  # SEM
            statistics.stdev,
            min,
            max
        ]
        for metric, func in zip(metrics, stats_funcs):
            row = [metric]
            for ratings in ratings_data:
                if ratings:
                    if func == statistics.stdev and len(ratings) < 2:
                        value = "N/A"
                    else:
                        try:
                            result = func(ratings)
                            # Check if result is already a string (like for CI function)
                            if isinstance(result, str):
                                value = result
                            else:
                                value = f"{result:.2f}"
                        except Exception:
                            value = "N/A"
                else:
                    value = "N/A"
                row.append(value)
            analysis_data.append(row)

        row = ["Sample Size (N)"]
        for ratings in ratings_data:
            row.append(str(len(ratings)) if ratings else "N/A")
        analysis_data.append(row)

        # 95% t-based CI half-width: t_{0.975, df=n-1} * s/sqrt(n)
        # If n < 2, "N/A"
        row = ["95% CI of Rating (T-based)"]
        for ratings in ratings_data:
            if ratings and len(ratings) > 1:
                n = len(ratings)
                m = statistics.mean(ratings)
                s = statistics.stdev(ratings)
                se = s / np.sqrt(n)
                t_crit = t.ppf(0.975, df=n-1)
                ci_half = t_crit * se
                value = f"{m:.2f} ±{ci_half:.2f}"
            elif ratings and len(ratings) == 1:
                # Only one rating, can't form a CI
                value = "N/A"
            else:
                value = "N/A"
            row.append(value)
        analysis_data.append(row)

    # Add total cost per chat
    row = ["Cost"]
    for chat_index in sorted(chat_results.keys()):
        cost = chat_results[chat_index]["total_cost"]
        row.append(f"${cost:.4f}")
    analysis_data.append(row)

    plot_base64_list = generate_plots(chat_results, analyze_length, analyze_rating)

    return analysis_data, plot_base64_list, total_cost

def generate_plots(chat_results, analyze_length, analyze_rating):
    plot_base64_list = []
    plt.style.use('default')
    plt.rcParams.update({'font.size': 16}) #bigger text for older eyes

    if analyze_length:
        plot_base64_list.extend(generate_length_plots(chat_results))

    if analyze_rating:
        plot_base64_list.extend(generate_rating_plots(chat_results))

        rating_violin = generate_rating_violin_plot(chat_results)
        if rating_violin is not None:
            plot_base64_list.append(rating_violin)

    return plot_base64_list

def generate_length_plots(chat_results):
    length_plots = []

    # Prepare data
    lengths_data = []
    labels = []
    for chat_index in sorted(chat_results.keys()):
        lengths = [l for l in chat_results[chat_index]["lengths"] if l is not None]
        if lengths:
            lengths_data.append(lengths)
            labels.append(f"Chat {chat_index}")

    # Check if there is data to plot
    if not lengths_data:
        return length_plots

    # Length distribution histogram
    fig_length_dist = plt.figure(figsize=(5, 6))
    ax = fig_length_dist.add_subplot(111)
    for lengths, label in zip(lengths_data, labels):
        hist, bins = np.histogram(lengths, bins=10, density=True)
        bin_centers = (bins[:-1] + bins[1:]) / 2
        ax.fill_between(bin_centers, hist, alpha=0.5, label=label)
    ax.set_title('Response Length Distribution')
    ax.set_xlabel('Response Length (characters)')
    ax.set_ylabel('Density')
    ax.legend()

    buf = io.BytesIO()
    plt.savefig(buf, format='png', bbox_inches='tight')
    buf.seek(0)
    length_plots.append(base64.b64encode(buf.getvalue()).decode('utf-8'))
    plt.close()

    # Mean length with 95% CI
    means = []
    cis = []
    for lengths in lengths_data:
        if len(lengths) > 1:
            m = statistics.mean(lengths)
            se = statistics.stdev(lengths) / np.sqrt(len(lengths))
            ci = 1.96 * se
            means.append(m)
            cis.append(ci)
        else:
            # If only one length, no CI can be computed
            means.append(lengths[0])
            cis.append(0)

    fig_length_ci = plt.figure(figsize=(5,6))
    ax = fig_length_ci.add_subplot(111)

    x_positions = np.arange(len(labels))
    ax.bar(x_positions, means, yerr=cis, align='center', alpha=0.7, ecolor='black', capsize=5)
    ax.set_title('Mean Response Length with 95% CI')
    ax.set_xticks(x_positions)
    ax.set_xticklabels(labels)
    ax.set_ylabel('Response Length (characters)')
    ax.yaxis.grid(True)

    buf = io.BytesIO()
    plt.savefig(buf, format='png', bbox_inches='tight')
    buf.seek(0)
    length_plots.append(base64.b64encode(buf.getvalue()).decode('utf-8'))
    plt.close()

    # Length boxplot
    fig_length_box = plt.figure(figsize=(5, 6))
    ax = fig_length_box.add_subplot(111)
    ax.boxplot(lengths_data, labels=labels)
    ax.set_title('Response Length Box Plot')
    ax.set_ylabel('Response Length (characters)')

    buf = io.BytesIO()
    plt.savefig(buf, format='png', bbox_inches='tight')
    buf.seek(0)
    length_plots.append(base64.b64encode(buf.getvalue()).decode('utf-8'))
    plt.close()

    return length_plots

def generate_rating_plots(chat_results):
    rating_plots = []

    # Prepare data
    ratings_data = []
    labels = []
    for chat_index in sorted(chat_results.keys()):
        ratings = [r for r in chat_results[chat_index]["ratings"] if r is not None]
        if ratings:
            ratings_data.append(ratings)
            labels.append(f"Chat {chat_index}")

    # Check if there is data to plot
    if not ratings_data:
        return rating_plots

    # Mean rating with 95% CI
    means = []
    cis = []
    for ratings in ratings_data:
        if len(ratings) > 1:
            m = statistics.mean(ratings)
            se = statistics.stdev(ratings) / np.sqrt(len(ratings))
            ci = 1.96 * se
            means.append(m)
            cis.append(ci)
        else:
            # If only one rating, no CI can be computed. We'll show no error bar.
            means.append(ratings[0])
            cis.append(0)

    fig_rating_ci = plt.figure(figsize=(5,6))
    ax = fig_rating_ci.add_subplot(111)

    x_positions = np.arange(len(labels))
    ax.bar(x_positions, means, yerr=cis, align='center', alpha=0.7, ecolor='black', capsize=5)
    ax.set_title('Mean Rating with 95% CI')
    ax.set_xticks(x_positions)
    ax.set_xticklabels(labels)
    ax.set_ylabel('Rating')
    ax.yaxis.grid(True)

    buf = io.BytesIO()
    plt.savefig(buf, format='png', bbox_inches='tight')
    buf.seek(0)
    rating_plots.append(base64.b64encode(buf.getvalue()).decode('utf-8'))
    plt.close()

    # Rating distribution histogram
    fig_rating_dist = plt.figure(figsize=(5, 6))
    ax = fig_rating_dist.add_subplot(111)

    max_rating = max(
        max(ratings) for ratings in ratings_data if ratings
    ) if ratings_data else 5

    bins = np.arange(0, int(max_rating) + 2)

    total_chats = len(ratings_data)
    bar_width = 0.8 / total_chats

    for idx, (ratings, label) in enumerate(zip(ratings_data, labels)):
        hist, _ = np.histogram(ratings, bins=bins)
        hist = hist / len(ratings)
        x = np.arange(len(bins) - 1)
        ax.bar(x + (idx * bar_width), hist, bar_width, label=label, alpha=0.7)

    ax.set_title('Rating Distribution')
    ax.set_xlabel('Rating')
    ax.set_ylabel('Proportion')
    ax.set_xticks(x + bar_width * (total_chats - 1) / 2)
    ax.set_xticklabels([f'{i:.1f}' for i in bins[:-1]])
    ax.legend()

    buf = io.BytesIO()
    plt.savefig(buf, format='png', bbox_inches='tight')
    buf.seek(0)
    rating_plots.append(base64.b64encode(buf.getvalue()).decode('utf-8'))
    plt.close()

    # Rating boxplot (to show distribution)
    fig_rating_box = plt.figure(figsize=(5, 6))
    ax = fig_rating_box.add_subplot(111)
    ax.boxplot(ratings_data, labels=labels)
    ax.set_title('Rating Box Plot')
    ax.set_ylabel('Rating')

    buf = io.BytesIO()
    plt.savefig(buf, format='png', bbox_inches='tight')
    buf.seek(0)
    rating_plots.append(base64.b64encode(buf.getvalue()).decode('utf-8'))
    plt.close()

    return rating_plots


def generate_rating_violin_plot(chat_results):
    # ADDED: A violin plot for ratings
    ratings_data = []
    labels = []
    for chat_index in sorted(chat_results.keys()):
        ratings = [r for r in chat_results[chat_index]["ratings"] if r is not None]
        if ratings:
            ratings_data.append(ratings)
            labels.append(f"Chat {chat_index}")

    if not ratings_data:
        return None

    fig_violin = plt.figure(figsize=(5,6))
    ax = fig_violin.add_subplot(111)
    ax.violinplot(ratings_data, showmeans=True, showextrema=True)
    ax.set_title('Rating Distribution (Violin Plot)')
    ax.set_xticks(np.arange(1, len(labels) + 1))
    ax.set_xticklabels(labels)
    ax.set_ylabel('Rating')

    buf = io.BytesIO()
    plt.savefig(buf, format='png', bbox_inches='tight')
    buf.seek(0)
    plot_base64 = base64.b64encode(buf.getvalue()).decode('utf-8')
    plt.close()
    return plot_base64


def create_html_report(
    analysis_data,
    plot_base64_list,
    total_cost,
    chat_data,
    chat_results,
    model_response=None,
    model_rating=None,
    temperature_response=None,
    temperature_rating=None,
    evaluation_rubrics=None,
    analyze_rating=True,
    show_transcripts=False,
):
    # Construct combined HTML for each chat's system message, prompts, and evaluation rubric
    def extract_user_messages(messages):
        conversation = []
        for msg in messages:
            if msg['role'] == 'user':
                conversation.append(f"&#x1F9D1; {msg['content']}")
            elif msg['role'] == 'assistant':
                # Show [AI Responds] for blank assistant messages
                content = msg['content'].strip()
                conversation.append(f"&#x1F916; {content if content else '[AI Responds]'}")
        conversation_html = ('<br>').join(conversation)
        return conversation_html

    combined_html = ""

    for idx, chat_info in enumerate(chat_data, start=1):
        rubric_html = ""
        if analyze_rating and evaluation_rubrics and idx in evaluation_rubrics:
            rubric = evaluation_rubrics[idx]
            rubric_html = f"""
                <div class='response-box'>
                    {rubric}
                </div>
            """
        
        combined_html += f"""
        <div class="config-section" style="margin-bottom: 0;">
            <p><strong>Chat {idx}</strong></p>
            <div class="config-columns">
                <div class="config-column">
                    <div class="prompt-label">System Message:</div>
                    <div class="response-box" style="margin-bottom: 0;">{chat_info['system_message'] if chat_info['system_message'] else 'None'}</div>
                </div>
                <div class="config-column">
                    <div class="prompt-label" style="margin-top: 0;">Prompts:</div>
                    <div class="response-box">{extract_user_messages([msg for msg in chat_info['messages'] if msg['role'] in ['user','assistant']])}</div>
                </div>
            </div>
                <div class="prompt-label">Evaluation Rubric:</div>
                {rubric_html}
            </div>
        """

    # CSS styles with escaped curly braces
    css_styles = """
        body { font-family: Arial, sans-serif; margin: 40px; }
        table { border-collapse: collapse; width: 100%; margin: 20px 0; }
        th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }
        th { background-color: #f2f2f2; }
        .response-box {
            background-color: #f8f9fa;
            border: 1px solid #ddd;
            padding: 15px;
            margin: 10px 0;
            border-radius: 4px;
        }
        .info-section {
            background-color: #f8f9fa;
            border: 1px solid #ddd;
            padding: 20px;
            margin: 20px 0;
            border-radius: 4px;
        }
        .config-columns {
            display: flex;
            gap: 20px;
            flex-wrap: wrap;
        }
        .config-column {
            flex: 1;
            min-width: 250px;
        }
        .prompt-container {
            margin: 20px 0;
        }
        .prompt-label {
            font-weight: bold;
            color: #555;
            margin-bottom: 5px;
            display: block;
        }
        .config-section {
            margin-bottom: 40px;
        }
        h1, h2, h3, h4, h5 {
            color: #333;
        }
    """

    # Build the HTML content using f-strings
    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <title>Test Results</title>
        <style>
            {css_styles}
        </style>
    </head>
    <body>
        <h1>Research Results</h1>

        <div class="info-section">
            <h2>Configuration</h2>
            <p><strong>Number of Iterations:</strong> {len(next(iter(chat_results.values()))['responses']) if chat_results else 'N/A'}</p>
            <p><strong>Response Generation Model:</strong> {model_response if model_response else 'N/A'}</p>
            <p><strong>Response Temperature:</strong> {temperature_response if temperature_response else 'N/A'}</p>
            {f"<p><strong>Rating Model:</strong> {model_rating}</p>" if analyze_rating and model_rating else ""}
            {f"<p><strong>Rating Temperature:</strong> {temperature_rating}</p>" if analyze_rating and temperature_rating else ""}
            {combined_html}
        </div>

        <h2>Cost Analysis</h2>
        <p>Total API cost: ${total_cost:.4f}</p>

        <h2>Detailed Analysis</h2>
        {tabulate(analysis_data, headers='firstrow', tablefmt='html')}

        {"".join([f"<h2>Visualizations</h2><img src='data:image/png;base64,{plot}' alt='Research Plot' style='max-width: 100%;'><br>" for plot in plot_base64_list]) if plot_base64_list else ""}
    """

    if show_transcripts and chat_results:
        transcripts_html = "<h2>Transcripts</h2>"

        # Sort chat_results by chat index to ensure Chat 1, Chat 2, etc.
        for chat_index in sorted(chat_results.keys()):
            results = chat_results[chat_index]
            transcripts_html += f"<h3>Chat {chat_index}</h3>"

            # Sort iterations to ensure Iteration 1, Iteration 2, etc.
            for iteration_idx, messages in enumerate(results["messages_per_iteration"], start=1):
                transcripts_html += f"<h4>Iteration {iteration_idx}</h4>"
                conversation_html = "<div class='response-box'>"
                for msg in messages:
                    role_symbol = '👤' if msg['role'] == 'user' else '🤖'
                    content = msg['content'].strip() if msg['content'].strip() else '[AI Responds]'
                    conversation_html += f"{role_symbol} {content}<br>"
                conversation_html += "</div>"
                transcripts_html += conversation_html

                if analyze_rating:
                    rating = results["ratings"][iteration_idx - 1] if results["ratings"] else "N/A"
                    rating_text = results["rating_texts"][iteration_idx - 1] if results["rating_texts"] else "N/A"
                    transcripts_html += f"""
                    <div class="response-box">&#x1F916; Rating: {rating} (raw response: "{rating_text}")</div>
                    """

        html_content += transcripts_html

    # Close the HTML tags
    html_content += """
    </body>
    </html>
    """

    return html_content


def generate_experiment_xlsx(
    settings_dict,
    chat_data,
    analysis_data,
    chat_results,
    plot_base64_list
):
    # Use the existing generate_settings_xlsx to create the initial BytesIO object
    initial_output = generate_settings_xlsx(settings_dict, chat_data, validate_chat=True)
    initial_output.seek(0)

    # Load the workbook from the BytesIO object using openpyxl
    workbook = load_workbook(initial_output)

    # Add Analysis Data sheet
    df_analysis = pd.DataFrame(analysis_data[1:], columns=analysis_data[0])
    sheet_analysis = workbook.create_sheet('Analysis Data')

    # Convert string values to numbers and apply formatting
    for r_idx, row in enumerate(dataframe_to_rows(df_analysis, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = sheet_analysis.cell(row=r_idx, column=c_idx, value=value)
                
            # Make header row boldface
            if r_idx == 1:
                cell.font = Font(bold=True)
                
            # Skip header row and first column (metric names)
            if r_idx > 1 and c_idx > 1:
                if isinstance(value, str):
                    if value == 'N/A':
                        cell.value = 'N/A'
                    elif value.startswith('$'):
                        # Convert currency string to number
                        cell.value = float(value.replace('$', ''))
                        cell.number_format = '"$"#,##0.0000'
                    else:
                        # Convert other numeric strings to numbers
                        try:
                            cell.value = float(value)
                            # Use 2 decimal places for ratings, 1 for lengths
                            if 'Rating' in df_analysis.iloc[r_idx-1, 0]:
                                cell.number_format = '0.00'
                            else:
                                cell.number_format = '0.0'
                        except ValueError:
                            # Keep as string if conversion fails
                            pass

    # Adjust column widths
    for column_cells in sheet_analysis.columns:
        length = max(len(str(cell.value) if cell.value is not None else "") for cell in column_cells)
        sheet_analysis.column_dimensions[column_cells[0].column_letter].width = length + 2

    # Create a consolidated Transcripts sheet
    sheet_transcripts = workbook.create_sheet('Transcripts')

    # Prepare data for transcripts
    columns_data = []
    max_messages = 0

    # Collect data per chat and iteration
    for chat_index in sorted(chat_results.keys()):
        chat_result = chat_results[chat_index]
        messages_per_iteration = chat_result['messages_per_iteration']
        ratings = chat_result.get('ratings', [])
        rating_texts = chat_result.get('rating_texts', [])

        for iteration_index, messages in enumerate(messages_per_iteration, 1):
            column = {
                'chat': f"Chat {chat_index}",
                'iteration': f"Iteration {iteration_index}",
                'messages': [],
                'rating_pos': None
            }

            # Collect messages
            for msg in messages:
                content = f"{msg['content']}"
                column['messages'].append(content)

            # Append rating at the end
            if ratings and iteration_index <= len(ratings):
                rating = ratings[iteration_index - 1]
                rating_text = rating_texts[iteration_index - 1] if rating_texts else ''
                rating_content = f"Rating: {rating} [verbatim rating: '{rating_text}']"
                column['rating_pos'] = len(column['messages']) + 2  # +2 to account for header rows
                column['messages'].append(rating_content)

            # Update max_messages
            if len(column['messages']) > max_messages:
                max_messages = len(column['messages'])

            columns_data.append(column)

    # Write headers
    sheet_transcripts.cell(row=1, column=1, value='Chat #')
    sheet_transcripts.cell(row=2, column=1, value='Iteration #')

    for col_idx, column in enumerate(columns_data, start=2):
        sheet_transcripts.cell(row=1, column=col_idx, value=column['chat'])
        sheet_transcripts.cell(row=2, column=col_idx, value=column['iteration'])

    # Fill messages and apply formatting
    for col_idx, column in enumerate(columns_data, start=2):
        for msg_idx, message in enumerate(column['messages'], start=3):
            cell = sheet_transcripts.cell(row=msg_idx, column=col_idx, value=message)
            cell.alignment = Alignment(wrap_text=True)

            # Boldface the rating cell
            if column['rating_pos'] and msg_idx == column['rating_pos']:
                cell.font = Font(bold=True)

    # Set column widths and wrap text
    sheet_transcripts.column_dimensions['A'].width = 15  # Width for the first column
    for col in sheet_transcripts.iter_cols(min_col=2, max_col=sheet_transcripts.max_column):
        col_letter = col[0].column_letter
        sheet_transcripts.column_dimensions[col_letter].width = 75
        for cell in col:
            cell.alignment = Alignment(wrap_text=True)

    # Modified plot section in Excel
    if plot_base64_list:
        img_sheet = workbook.create_sheet("Analysis Plots")
        
        # Define layout parameters
        row_height = 2  # Height in Excel rows
        images_per_row = 3
        
        for i, plot_base64 in enumerate(plot_base64_list):
            img_data = base64.b64decode(plot_base64)
            image_stream = io.BytesIO(img_data)
            img = OpenPyXLImage(image_stream)
            
            # Scale image
            scale_factor = 0.5
            img.width = img.width * scale_factor
            img.height = img.height * scale_factor
            
            # Calculate position
            row = (i // images_per_row) * row_height + 1
            col = (i % images_per_row) * 6 + 1  # 4 columns spacing between images
            
            # Add image to worksheet
            img_sheet.add_image(img, f'{get_column_letter(col)}{row}')
            
            # Adjust row height to accommodate images
            img_sheet.row_dimensions[row].height = 250

    # Save the workbook to a new BytesIO object
    final_output = io.BytesIO()
    workbook.save(final_output)
    final_output.seek(0)

    return final_output
